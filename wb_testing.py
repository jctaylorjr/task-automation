from collections import defaultdict
from openpyxl import load_workbook
from mapping import find_printer_tasks

wb = load_workbook("MGH_Label Printer Data Collection Sheet_Logical Mapping-testing2.xlsx")
sheet_ranges = wb["Data Collection Sheet"]


def find_printer_tasks(control_id_column, task_column, sheet_ranges):
    printer_tasks = defaultdict(list)

    for row in range(2, sheet_ranges.max_row + 1):
        control_id = sheet_ranges[f"{control_id_column}{row}"].value
        task = sheet_ranges[f"{task_column}{row}"].value
        if control_id is None or task is None:
            continue
        printer_tasks[control_id] = task

    return printer_tasks


# print(type(sheet_ranges["P1323"]))
# print(type(sheet_ranges["P1323"].value))
# print(sheet_ranges["P1323"].value)

# tasks = find_printer_tasks("C", "P", sheet_ranges)
# print(tasks[736720])

# mappings = printer_workstation_mapping("C", "B", sheet_ranges)
# print(mappings[738213])
# print_workstations(mappings[738213])
# print(printer_to_workstations.get(702772))
# for mapping, values in mappings.items():
#     print(f"{mapping}({len(values)}): {" ".join(values)}")
    # print_workstations(values)

# seen = seen_printer("C", "B", sheet_ranges)
# print(seen)

# for cid in seen:
#     print(cid)
#     print_workstations(mappings[cid])



# with sync_playwright() as playwright:
#     chromium = playwright.chromium # or "firefox" or "webkit".
#     browser = chromium.launch(headless=True)
#     context = browser.new_context()
#     ws_page = context.new_page()
#     ws_page.goto("https://partnershealthcare.service-now.com/esp?id=esp_workstations")
#     new_ws_list = get_workstation_lwsids(ws_page, [293270, 210896])
#     print(", ".join(new_ws_list))


# def get_workstation_lwsids(page: Page, ws_names):
#         print("Looking up workstation LWSIDs...")
#         ws_name_lwsids = []
#         # for _ in range(5):
#         #     try:
#         #         page.goto("https://partnershealthcare.service-now.com/esp?id=esp_workstations")
#         #     except TimeoutError:
#         #         print("Retry goto...")
#         #         continue
#         #     else:
#         #         break
#         for ws in ws_names:
#             for i in range(5):
#                 try:
#                     if i > 0:
#                         print(f"Retrying {ws}...")
#                     page.locator("#inputTxtWksta").fill(ws)
#                     with page.expect_response("https://partnershealthcare.service-now.com/api/now/sp/rectangle/ddd076c2473b39d02a94f147536d4327?id=esp_workstations") as response_info:
#                         page.locator("#wkstaSearch").click()
#                 except TimeoutError:
#                     continue
#                 else:
#                     response = response_info.value
#                     if response.ok:
#                         try:
#                             ws_name_lwsid = response.json()["result"]["data"]["lws"]["searchList"][0]["displayValue"]
#                         except:
#                             ws_name_lwsid = f"{ws} [Not found]"
#                             continue
#                 finally:
#                     ws_name_lwsids.append(ws_name_lwsid)
#         return ws_name_lwsids
