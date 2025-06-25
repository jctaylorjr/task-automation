from openpyxl import load_workbook
from collections import defaultdict
from playwright.sync_api import sync_playwright, Playwright, Page
import math

def print_workstations(workstations):
    if len(workstations) > 1:
        print(" ".join(workstations))

def printer_workstation_mapping(control_id_column, workstation_column, sheet_ranges):
    printer_to_workstations = defaultdict(list)

    for row in range(2, sheet_ranges.max_row + 1):
        control_id = sheet_ranges[f"{control_id_column}{row}"].value
        workstation = sheet_ranges[f"{workstation_column}{row}"].value
        if control_id is None or workstation is None:
            continue
        printer_to_workstations[control_id].append(str(workstation))

    return printer_to_workstations

def get_workstation_lwsids(page: Page, ws_names):
    print("Looking up workstation LWSIDs...")
    ws_name_lwsids = []
    for ws in ws_names:
        if type(ws) is int:
            ws = f"W0{str(ws)}"
        # to make sure W0253604 and DIMFLH2664 can get in
        if len(ws) != 8 and len(ws) != 10:
            continue
        page.locator("#inputTxtWksta").fill(ws)
        with page.expect_response("https://partnershealthcare.service-now.com/api/now/sp/rectangle/ddd076c2473b39d02a94f147536d4327?id=esp_workstations") as response_info:
            page.locator("#wkstaSearch").click()
            response = response_info.value
            if response.ok:
                try:
                    ws_name_lwsid = response.json()["result"]["data"]["lws"]["searchList"][0]["displayValue"]
                except:
                    ws_name_lwsid = f"{ws} [Not found]"
                finally:
                    ws_name_lwsids.append(ws_name_lwsid)
    return ws_name_lwsids

def seen_printer(control_id_column, task_column, sheet_ranges):
    seen_printers = set()

    for row in range(2, sheet_ranges.max_row + 1):
        control_id = sheet_ranges[f"{control_id_column}{row}"].value
        task = sheet_ranges[f"{task_column}{row}"].value
        if control_id is not None and task is not None and control_id not in seen_printers:
            seen_printers.add(control_id)
        
    return seen_printers

# wb = load_workbook("MGH_Label Printer Data Collection Sheet_Logical Mapping-testing.xlsx")
# sheet_ranges = wb["Data Collection Sheet"]
# mappings = printer_workstation_mapping("C", "B", sheet_ranges)
# # print(mappings[738213])
# # print_workstations(mappings[738213])
# # print(printer_to_workstations.get(702772))
# for mapping, values in mappings.items():
#     print(f"{mapping}({len(values)}): {" ".join(values)}")
#     # print_workstations(values)

# seen = seen_printer("C", "B", sheet_ranges)
# print(seen)

# for cid in seen:
#     print(cid)
#     print_workstations(mappings[cid])



# with sync_playwright() as playwright:
#     chromium = playwright.chromium # or "firefox" or "webkit".
#     browser = chromium.launch(headless=True)
#     context = browser.new_context()
#     ws_page = context.new_page()
#     ws_page.goto("https://partnershealthcare.service-now.com/esp?id=esp_workstations")
#     new_ws_list = get_workstation_lwsids(ws_page, [293270, 210896])
#     print(", ".join(new_ws_list))


# def get_workstation_lwsids(page: Page, ws_names):
#         print("Looking up workstation LWSIDs...")
#         ws_name_lwsids = []
#         # for _ in range(5):
#         #     try:
#         #         page.goto("https://partnershealthcare.service-now.com/esp?id=esp_workstations")
#         #     except TimeoutError:
#         #         print("Retry goto...")
#         #         continue
#         #     else:
#         #         break
#         for ws in ws_names:
#             for i in range(5):
#                 try:
#                     if i > 0:
#                         print(f"Retrying {ws}...")
#                     page.locator("#inputTxtWksta").fill(ws)
#                     with page.expect_response("https://partnershealthcare.service-now.com/api/now/sp/rectangle/ddd076c2473b39d02a94f147536d4327?id=esp_workstations") as response_info:
#                         page.locator("#wkstaSearch").click()
#                 except TimeoutError:
#                     continue
#                 else:
#                     response = response_info.value
#                     if response.ok:
#                         try:
#                             ws_name_lwsid = response.json()["result"]["data"]["lws"]["searchList"][0]["displayValue"]
#                         except:
#                             ws_name_lwsid = f"{ws} [Not found]"
#                             continue
#                 finally:
#                     ws_name_lwsids.append(ws_name_lwsid)
#         return ws_name_lwsids
